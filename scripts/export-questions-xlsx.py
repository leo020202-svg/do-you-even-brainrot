"""Export data/questions.json to a clean .xlsx workbook.

Sheet 1 "Questions": one row per question with formula-driven correct-text lookup.
Sheet 2 "Summary":   COUNTIF-driven totals by category and difficulty.
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "data" / "questions.json"
OUT = ROOT / "data" / "questions.xlsx"

HEADER_FILL = PatternFill("solid", start_color="0F0A1E")
HEADER_FONT = Font(name="Arial", bold=True, color="A8FF3E", size=11)
BODY_FONT = Font(name="Arial", size=11)
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")

COLUMNS = [
    ("id", 8),
    ("category", 18),
    ("difficulty", 11),
    ("question", 60),
    ("option_a", 36),
    ("option_b", 36),
    ("option_c", 36),
    ("option_d", 36),
    ("correct_answer", 9),
    ("correct_text", 40),
    ("source", 22),
    ("tags", 18),
    ("active", 8),
]


def load() -> list[dict]:
    with SRC.open("r", encoding="utf-8") as f:
        return json.load(f)


def build_questions_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.active
    ws.title = "Questions"

    headers = [c[0] for c in COLUMNS]
    ws.append(headers)
    for col_idx, (_, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", horizontal="left")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for i, q in enumerate(rows, start=2):
        opts = {o["id"]: o["text"] for o in q.get("options", [])}
        ws.cell(row=i, column=1, value=q.get("id", ""))
        ws.cell(row=i, column=2, value=q.get("category", ""))
        ws.cell(row=i, column=3, value=q.get("difficulty", ""))
        ws.cell(row=i, column=4, value=q.get("question", ""))
        ws.cell(row=i, column=5, value=opts.get("A", ""))
        ws.cell(row=i, column=6, value=opts.get("B", ""))
        ws.cell(row=i, column=7, value=opts.get("C", ""))
        ws.cell(row=i, column=8, value=opts.get("D", ""))
        ws.cell(row=i, column=9, value=q.get("correct_answer", ""))
        # Formula picks the correct option's text from E:H based on the letter in I.
        ws.cell(
            row=i,
            column=10,
            value=f'=CHOOSE(MATCH(I{i},{{"A","B","C","D"}},0),E{i},F{i},G{i},H{i})',
        )
        ws.cell(row=i, column=11, value=q.get("source") or "")
        ws.cell(row=i, column=12, value=", ".join(q.get("tags", []) or []))
        ws.cell(row=i, column=13, value=bool(q.get("active", True)))

        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=i, column=col_idx)
            cell.font = BODY_FONT
            cell.alignment = WRAP if col_idx in (4, 5, 6, 7, 8, 10) else TOP

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(COLUMNS))
    last_row = len(rows) + 1
    table = Table(displayName="Questions", ref=f"A1:{last_col}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)


def build_summary_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14

    def header(cell: str, text: str) -> None:
        c = ws[cell]
        c.value = text
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    header("A1", "Category")
    header("B1", "Count")
    header("D1", "Difficulty")
    header("E1", "Count")

    categories = sorted({q.get("category", "") for q in rows if q.get("category")})
    for i, cat in enumerate(categories, start=2):
        ws.cell(row=i, column=1, value=cat).font = BODY_FONT
        ws.cell(row=i, column=2, value=f'=COUNTIF(Questions!B:B,A{i})').font = BODY_FONT

    total_row = len(categories) + 2
    ws.cell(row=total_row, column=1, value="Total").font = Font(name="Arial", bold=True)
    ws.cell(
        row=total_row, column=2, value=f"=SUM(B2:B{total_row - 1})"
    ).font = Font(name="Arial", bold=True)

    difficulties = ["easy", "medium", "hard", "expert"]
    for i, diff in enumerate(difficulties, start=2):
        ws.cell(row=i, column=4, value=diff).font = BODY_FONT
        ws.cell(row=i, column=5, value=f'=COUNTIF(Questions!C:C,D{i})').font = BODY_FONT

    diff_total_row = len(difficulties) + 2
    ws.cell(row=diff_total_row, column=4, value="Total").font = Font(name="Arial", bold=True)
    ws.cell(
        row=diff_total_row, column=5, value=f"=SUM(E2:E{diff_total_row - 1})"
    ).font = Font(name="Arial", bold=True)


def main() -> None:
    rows = load()
    wb = Workbook()
    build_questions_sheet(wb, rows)
    build_summary_sheet(wb, rows)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT} ({len(rows)} questions)")


if __name__ == "__main__":
    main()
